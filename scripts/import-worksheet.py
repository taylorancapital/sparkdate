#!/usr/bin/env python3
"""
scripts/import-worksheet.py

ONE-TIME conversion: SparkDate_Posting_Worksheet.xlsx -> content/queue.csv.

Why Python for this and Node for everything else: this runs once, and openpyxl
is already installed and proven (the posting skill uses it). Adding an xlsx
dependency to package.json for a single conversion would put supply-chain
surface in a repo that deliberately carries four dependencies and no build
step. Everything that runs on an ongoing basis -- the linter, the publishers --
is Node against lib/content-queue.js, which uses no dependency at all.

After this runs, content/queue.csv is the source of truth. The six xlsx copies
in Downloads/ should be archived; leaving them live is how you get a seventh.

Usage:
  python scripts/import-worksheet.py                      # default source path
  python scripts/import-worksheet.py --src path/to.xlsx
  python scripts/import-worksheet.py --dry-run            # print, write nothing
"""

import argparse
import csv
import datetime
import glob
import json
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)

DEFAULT_SRC = os.path.expanduser(
    r"~/Downloads/SparkDate_Posting_Worksheet_2.xlsx"
)
# Where the already-exported PNGs live. Used only to pre-fill asset_files;
# a miss is a warning, never fatal -- Layer 2 renames them anyway.
DEFAULT_ASSET_DIR = os.path.expanduser(
    r"~/Downloads/30 Day Content Plan Claude Posting"
)

OUT = os.path.join(REPO, "content", "queue.csv")
BRAND = os.path.join(REPO, "content", "brand.json")

# Worksheet event names -> brand.json event keys. A row can name more than one
# event (merged recap-plus-launch posts), which is why `events` is a list:
# the linter accepts hashtags from the UNION of the named events' pools, so
# GG-7's Lancaster tags on a Philly-recap row validate without a special case.
EVENT_MAP = {
    "Tellus AfterDark (Tellus360)": ["TL"],
    "Marion Court": ["MC"],
    "Good Good Things": ["GG"],
    "Loxley's": ["LX"],
    "Good Good Things + Loxley's": ["GG", "LX"],
    "Marion Court + Loxley's": ["MC", "LX"],
}

STATE_MAP = {
    "POSTED": "posted",
    "DARK": "skipped",
    "SCHEDULE": "pending",
    "STORY": "pending",
    "CONFLICT": "pending",
}

COLUMNS = [
    "row_id", "date", "time", "events", "platforms", "format", "state",
    "caption", "hashtags", "caption_x", "link_fb", "link_ig",
    "utm_campaign", "utm_content", "asset_files", "manual_reason", "notes",
]


def pad_row_id(raw):
    """MC-1 -> MC-01. Without zero-padding MC-10 sorts before MC-2 and the
    queue falls out of posting order -- the brand doc calls this out."""
    s = str(raw).strip()
    m = re.match(r"^([A-Z]+)-(\d+)$", s)
    if m:
        return "%s-%02d" % (m.group(1), int(m.group(2)))
    return s  # MC-LAUNCH and friends pass through unchanged


def parse_time_24h(raw):
    if not raw:
        return ""
    s = str(raw).strip()
    for fmt in ("%I:%M %p", "%I %p", "%H:%M"):
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%H:%M")
        except ValueError:
            pass
    return s


def parse_platforms(raw, fmt):
    """'IG + FB' -> ['ig','fb']. Story formats get a story marker so the
    publisher knows not to try the feed endpoint."""
    s = (raw or "").lower()
    out = []
    if "ig" in s or "instagram" in s:
        out.append("ig")
    if "fb" in s or "facebook" in s:
        out.append("fb")
    if "tiktok" in s:
        out.append("tiktok")
    f = (fmt or "").lower()
    if "story" in f and "ig" in out:
        out = [p for p in out if p != "ig"] + ["ig_story"]
    return out


def extract_links(raw):
    """Pull the Facebook and Instagram URLs out of the free-text Link handling
    cell. They are distinguished by utm_source, not by position."""
    if not raw:
        return "", ""
    urls = re.findall(r"https?://[^\s\"'<>)\]]+", str(raw))
    fb = ig = ""
    for u in urls:
        u = u.rstrip(".,;")
        if "utm_source=Facebook" in u and not fb:
            fb = u
        elif "utm_source=Instagram" in u and not ig:
            ig = u
    return fb, ig


def rewrite_utms(url, row_id, event_keys, date_str):
    """Replace the shared utm_content=proof_rsa1 with the row id, and the
    stale Augweek3_* campaign with {event}_{YYYYMM}. This is what makes GA4
    able to attribute an individual post instead of only the channel."""
    if not url:
        return "", "", ""
    ym = (date_str or "").replace("-", "")[:6]
    campaign = "%s_%s" % ("-".join(event_keys), ym) if event_keys and ym else ""
    out = url
    if campaign:
        out = re.sub(r"utm_campaign=[^&\s]*", "utm_campaign=" + campaign, out)
    out = re.sub(r"utm_content=[^&\s]*", "utm_content=" + row_id, out)
    return out, campaign, row_id


def index_assets(asset_dir):
    """Map row id -> ordered slide filenames, from names like
    sparkdate-aug23-1of3-MC2.png / tellus-aug19-3of3 -TL1.png (note the stray
    space -- match tolerantly rather than assume the exports are tidy)."""
    idx = {}
    if not os.path.isdir(asset_dir):
        return idx
    for path in glob.glob(os.path.join(asset_dir, "*.png")):
        base = os.path.basename(path)
        m = re.search(r"-(\d+)of(\d+)\s*-\s*([A-Za-z]+\d*)\s*\.png$", base)
        if not m:
            continue
        n, _total, tag = int(m.group(1)), m.group(2), m.group(3)
        tag = tag.upper()
        tm = re.match(r"^([A-Z]+)(\d+)$", tag)
        key = "%s-%02d" % (tm.group(1), int(tm.group(2))) if tm else (
            "MC-LAUNCH" if tag == "LAUNCH" else tag
        )
        idx.setdefault(key, []).append((n, base))
    return {k: [b for _, b in sorted(v)] for k, v in idx.items()}


def derive_caption_x(caption, link):
    """X posts cap at 280 and every link costs a flat 23 (t.co), so the text
    must fit in ~250. Trim on a sentence boundary rather than mid-word, and
    report anything that lost content so a human reviews it rather than
    shipping a truncated thought."""
    if not caption:
        return "", False
    flat = re.sub(r"\n+", " ", str(caption)).strip()
    budget = 280 - 24 if link else 280
    if len(flat) <= budget:
        return flat, False
    cut = flat[:budget]
    for sep in (". ", "! ", "? "):
        i = cut.rfind(sep)
        if i > budget * 0.5:
            return cut[: i + 1].strip(), True
    i = cut.rfind(" ")
    return (cut[:i] if i > 0 else cut).strip(), True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=DEFAULT_SRC)
    ap.add_argument("--assets", default=DEFAULT_ASSET_DIR)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        sys.exit("ERROR: openpyxl not installed. pip install openpyxl")

    if not os.path.exists(args.src):
        sys.exit("ERROR: source workbook not found: %s" % args.src)

    with open(BRAND, encoding="utf-8") as fh:
        brand = json.load(fh)

    assets = index_assets(args.assets)
    print("indexed assets for %d rows from %s" % (len(assets), args.assets))

    wb = openpyxl.load_workbook(args.src)
    ws = wb["Schedule"]
    hdr = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(hdr)}

    rows, warnings = [], []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if raw[0] is None:
            continue
        get = lambda name: raw[col[name]] if name in col else None

        row_id = pad_row_id(get("#"))
        date_str = str(get("Date") or "").strip()[:10]
        fmt = str(get("Format") or "").strip()
        ev_raw = str(get("Event") or "").strip()
        events = EVENT_MAP.get(ev_raw)
        if events is None:
            warnings.append("%s: unmapped event name %r" % (row_id, ev_raw))
            events = []

        platforms = parse_platforms(get("Platform"), fmt)
        fb, ig = extract_links(get("Link handling"))
        fb, campaign, content = rewrite_utms(fb, row_id, events, date_str)
        ig, campaign2, _ = rewrite_utms(ig, row_id, events, date_str)
        campaign = campaign or campaign2

        caption = str(get("Caption (paste verbatim)") or "")
        cap_x, trimmed = derive_caption_x(caption, fb or ig)
        if trimmed:
            warnings.append("%s: caption_x was trimmed to fit 280 -- review it" % row_id)

        state = STATE_MAP.get(str(get("Status") or "").strip().upper(), "pending")

        manual = ""
        if "story" in fmt.lower() and ("link sticker" in fmt.lower() or "countdown" in fmt.lower()):
            manual = "IG Story stickers (link/countdown) cannot be set via API"

        files = assets.get(row_id, [])
        if not files and state == "pending" and "NO POST" not in fmt.upper():
            warnings.append("%s: no asset files matched" % row_id)

        rows.append({
            "row_id": row_id,
            "date": date_str,
            "time": parse_time_24h(get("Time")),
            "events": ",".join(events),
            "platforms": ",".join(platforms),
            "format": fmt,
            "state": state,
            "caption": caption,
            "hashtags": str(get("Hashtags") or "").strip(),
            "caption_x": cap_x,
            "link_fb": fb,
            "link_ig": ig,
            "utm_campaign": campaign,
            "utm_content": content or row_id,
            "asset_files": ",".join(files),
            "manual_reason": manual,
            "notes": str(get("Caption source / notes") or "").strip(),
        })

    rows.sort(key=lambda r: (r["date"], r["time"], r["row_id"]))

    print("\nparsed %d rows" % len(rows))
    from collections import Counter
    print("state:", dict(Counter(r["state"] for r in rows)))
    print("with assets: %d / %d" % (sum(1 for r in rows if r["asset_files"]), len(rows)))
    if warnings:
        print("\n%d warnings:" % len(warnings))
        for w in warnings:
            print("  !", w)

    if args.dry_run:
        print("\n--dry-run: nothing written")
        return

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=COLUMNS, lineterminator="\n")
        w.writeheader()
        w.writerows(rows)
    print("\nwrote %s (%d rows)" % (OUT, len(rows)))


if __name__ == "__main__":
    main()
