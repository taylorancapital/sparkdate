#!/usr/bin/env python3
"""
scripts/sync-utm-content.py

Reads every servable Meta ad and writes a "Paid Ad UTMs" sheet mapping each
utm_content value to the ad, ad set, campaign and creative that carries it.

An ad tags itself in one of two places -- parameters baked into the destination
link, or the creative's `url_tags`, which Meta appends at click time -- and this
reads BOTH, reporting which one an ad uses in the `tagged_via` column. Reading
only the link (the behaviour until 2026-08-30) inverted the whole sheet: ads
tagged the correct way came back "(none)" while ads tagged the wrong way listed
perfectly. See effective() for the full account.

WHY THIS EXISTS

Until 2026-08-29 every ad in the account carried utm_content=proof_rsa1, so GA4
could not tell one creative from another -- three consecutive nightly reports
said per-ad attribution was impossible. Now that ads carry distinct values, the
mapping from a GA4 row back to "which ad, which video" lives only in Ads
Manager, which is exactly the sort of thing that drifts the moment nobody
maintains it by hand.

WHAT IT DOES NOT DO

It does NOT touch the existing "UTM Links" sheet. That sheet is hand-maintained
and holds organic/manual links; overwriting rows a human curated is how you lose
work you cannot get back. This writes a SEPARATE sheet, regenerated whole on
every run, and every value in it comes from the Meta API rather than being typed.

If the sheet does not exist it is created. If it does, its contents are cleared
and rewritten -- safe precisely because nothing in it is hand-entered.

Env:
  META_ADS_ACCESS_TOKEN   ads_read is sufficient (no writes)
  META_AD_ACCOUNT_ID      optional -- discovered if unset

Usage:
  python scripts/sync-utm-content.py --dry-run     # print, write nothing
  python scripts/sync-utm-content.py
  python scripts/sync-utm-content.py --workbook "path/to/other.xlsx"
"""

import argparse
import datetime as _dt
import json
import os
import sys
import urllib.parse
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)

# The UTM home is a FOLDER, not a file.
#
# This used to write into "Marketing & GTM/Content Calendar & UTM Links.xlsb
# .xlsx" -- a content calendar (Instagram/TikTok calendars, Daily Tracker,
# Dashboard Results) that had UTM sheets bolted onto it. The purpose-built
# workbook is "SparkDate_UTM_Campaign_Links", one sheet per event with an
# auto-assembling Full Campaign URL column, and it lives in a different
# folder entirely.
#
# The generated sheet does NOT go into that workbook, for a measured reason.
# openpyxl cannot round-trip x14 extended data validations, and the campaign
# workbook has 16 of them: the Platform dropdowns on column A of all four
# event sheets, sourced from Sheet2. A load+save preserves its 138 formulas
# and its inline utm_medium dropdown but silently destroys those Platform
# dropdowns -- verified on a copy, 16 elements in, 0 out. Re-injecting them
# means splicing raw XML and re-declaring namespace prefixes on openpyxl's
# output root, which is not a dependency worth carrying to keep two kinds of
# data in one file.
#
# So: script-owned output lives in its own file, in the SAME folder as the
# hand-maintained one. One place to look, nothing hand-curated at risk.
UTM_DIR = os.path.join(
    REPO, "Business Plan", "files", "Social Media Marketing",
    "Content Calendars & Strategy",
)
DEFAULT_WB = os.path.join(UTM_DIR, "SparkDate_Paid_Ad_UTMs (generated).xlsx")

# The hand-maintained companion, named here only so the banner can point at it.
HAND_WB = "SparkDate_UTM_Campaign_Links (version 1).xlsb.xlsx"

SHEET = "Paid Ad UTMs"
GRAPH = "https://graph.facebook.com/v21.0"

# Anything not in here can still serve. IN_PROCESS and PENDING_REVIEW are live:
# a filter on ACTIVE alone hides the ad you just built, which is when you look.
DORMANT = {"PAUSED", "ADSET_PAUSED", "CAMPAIGN_PAUSED", "ARCHIVED", "DELETED", "DISAPPROVED"}


def api(path, token, **params):
    params["access_token"] = token
    url = f"{GRAPH}/{path}?{urllib.parse.urlencode(params)}"
    with urllib.request.urlopen(url, timeout=60) as r:
        body = json.load(r)
    if "error" in body:
        raise SystemExit(f"Meta API: {body['error'].get('message')}")
    return body


def account_id(token):
    if os.environ.get("META_AD_ACCOUNT_ID"):
        return os.environ["META_AD_ACCOUNT_ID"]
    data = api("me/adaccounts", token, fields="id,name", limit=50).get("data", [])
    if len(data) == 1:
        return data[0]["id"]
    raise SystemExit(f"{len(data)} ad accounts visible -- set META_AD_ACCOUNT_ID.")


def qparam(query, key):
    """Read one key from a BARE query string, tolerating {{dynamic}} placeholders."""
    if not query:
        return ""
    q = urllib.parse.parse_qs(query, keep_blank_values=True)
    return (q.get(key) or [""])[0]


def param(url, key):
    """Read one query parameter from a full URL."""
    if not url or "?" not in url:
        return ""
    return qparam(url.split("?", 1)[1], key)


def effective(link, tags, key):
    """
    What our code actually records for `key`, given BOTH places an ad can
    carry it.

    An ad tags itself in one of two ways: parameters baked into the
    destination link, or the creative's `url_tags`, which Meta APPENDS to
    that link at click time. This script read only the link until
    2026-08-30, so the ads tagged the CORRECT way -- url_tags, with a clean
    destination -- came back as "(none)" and were even reported as colliding
    with each other, while the six proof_rsa1 ads tagged the wrong way
    listed perfectly. The registry that exists to trace a GA4 row back to an
    ad was blind to exactly the convention brand.json mandates, so every
    correctly-built ad from here on would have been invisible to it.

    Link wins when both carry the key: Meta appends, so the parameter appears
    twice and the link's copy comes first -- and first is the one our code
    reads. That is a defect, not a preference (see `tagged_via` == "both"),
    but the sheet must report what an ad ACTUALLY sends, not what it should.
    """
    return param(link, key) or qparam(tags, key)


def tagged_via(link, tags):
    """Where an ad's utm_source lives -- the field that made this bug invisible."""
    in_link = bool(param(link, "utm_source"))
    in_tags = bool(qparam(tags, "utm_source"))
    if in_link and in_tags:
        return "both"          # sends two utm_source values; the link's wins
    if in_link:
        return "link"
    if in_tags:
        return "url_tags"
    return "none"


def collect(token, act):
    fields = (
        "id,name,status,effective_status,updated_time,"
        "campaign{name},adset{name},"
        # url_tags is half of how an ad tags itself. Omitting it is what made
        # the two correctly-tagged Loxleys ads read as untagged.
        "creative{id,name,video_id,image_hash,url_tags,object_story_spec}"
    )
    rows = []
    resp = api(f"{act}/ads", token, fields=fields, limit=200)
    for ad in resp.get("data", []):
        cre = ad.get("creative") or {}
        spec = cre.get("object_story_spec") or {}
        data = spec.get("video_data") or spec.get("link_data") or {}
        cta = data.get("call_to_action") or {}
        link = (cta.get("value") or {}).get("link") or data.get("link") or ""
        tags = cre.get("url_tags") or ""
        media = "video" if spec.get("video_data") else ("image" if data.get("image_hash") else "none")
        rows.append({
            "utm_content": effective(link, tags, "utm_content"),
            "utm_source": effective(link, tags, "utm_source"),
            "utm_campaign": effective(link, tags, "utm_campaign"),
            "tagged_via": tagged_via(link, tags),
            "url_tags": tags,
            "ad": ad.get("name", ""),
            "adset": (ad.get("adset") or {}).get("name", ""),
            "campaign": (ad.get("campaign") or {}).get("name", ""),
            "media": media,
            "video_id": data.get("video_id") or cre.get("video_id") or "",
            "creative": cre.get("name", ""),
            "status": ad.get("effective_status", ""),
            "serving": "yes" if ad.get("effective_status") not in DORMANT else "no",
            "eventId": param(link, "eventId"),
            "url": link,
            "ad_id": ad.get("id", ""),
            "updated": ad.get("updated_time", ""),
        })
    rows.sort(key=lambda r: (r["serving"] != "yes", r["campaign"], r["ad"]))
    return rows


COLUMNS = [
    ("utm_content", 26), ("tagged_via", 11), ("serving", 9), ("media", 8),
    ("ad", 34), ("adset", 30), ("campaign", 34), ("creative", 30),
    ("utm_source", 22), ("utm_campaign", 22), ("eventId", 24), ("video_id", 20),
    ("status", 16), ("ad_id", 20), ("updated", 22), ("url", 60), ("url_tags", 70),
]


def has_extended_validations(path):
    """
    Count x14 (extended) data validations in a workbook.

    openpyxl silently drops these on save. They are how Excel expresses a
    dropdown whose list lives on ANOTHER sheet, so they are exactly the
    validations a hand-maintained workbook depends on. Anything above zero
    means saving with openpyxl would destroy work.
    """
    import re
    import zipfile
    with zipfile.ZipFile(path) as z:
        return sum(
            len(re.findall(r"x14:dataValidation\b", z.read(n).decode("utf-8", "ignore")))
            for n in z.namelist() if n.startswith("xl/worksheets/sheet")
        )


def write_sheet(path, rows):
    import openpyxl
    from openpyxl.styles import Font

    if os.path.exists(path):
        # Refuse rather than quietly degrade someone's workbook. This is the
        # guard that stops a future --workbook pointed at the hand-maintained
        # campaign links file from stripping its Platform dropdowns.
        lost = has_extended_validations(path)
        if lost:
            raise SystemExit(
                f"Refusing to write {os.path.basename(path)}: it carries {lost} extended "
                "(x14) data validation(s) -- dropdowns whose list lives on another sheet.\n"
                "openpyxl cannot round-trip those; saving would delete them permanently.\n"
                "Point --workbook at a script-owned file instead, and keep hand-maintained "
                "workbooks out of this script's reach."
            )
        wb = openpyxl.load_workbook(path)
    else:
        # First run: the generated workbook is ours to create.
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    if SHEET in wb.sheetnames:
        del wb[SHEET]          # regenerated whole; nothing here is hand-entered
    ws = wb.create_sheet(SHEET)

    stamp = _dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws.cell(row=1, column=1, value=f"Generated from the Meta Ads API {stamp} by scripts/sync-utm-content.py "
                                   f"-- do not hand-edit, it is overwritten on every run. "
                                   f"Hand-maintained links (Eventbrite, email, flyers, Nextdoor, Google Business) "
                                   f"belong in '{HAND_WB}' in this folder, which no script touches.")
    ws.cell(row=1, column=1).font = Font(italic=True, size=9)

    # The convention, next to the data it governs -- utm_content is REQUIRED
    # and unique per ad, not the "optional A/B field" the older guidance said.
    # A shared value is why 11 ads carried proof_rsa1 and none of them could
    # be attributed. See content/brand.json paid_template.caption_rules.utm.
    ws.cell(row=2, column=1, value="Convention: utm_source={{site_source_name}} (Meta fills it per placement, lowercase) "
                                   "| utm_medium=paid_social | utm_campaign={EVENT}_{YYYYMM} "
                                   "| utm_content={event}_{phase}_{ad_set}_{creative} -- REQUIRED, and unique per ad.")
    ws.cell(row=2, column=1).font = Font(italic=True, size=9)

    # Row 1 banner, row 2 convention, row 3 header, data from row 4.
    for c, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=3, column=c, value=name)
        cell.font = Font(bold=True)
        ws.column_dimensions[cell.column_letter].width = width
    for r, row in enumerate(rows, 4):
        for c, (name, _) in enumerate(COLUMNS, 1):
            ws.cell(row=r, column=c, value=row.get(name, ""))
    ws.freeze_panes = "A4"
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--workbook", default=DEFAULT_WB)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    token = os.environ.get("META_ADS_ACCESS_TOKEN")
    if not token:
        raise SystemExit("META_ADS_ACCESS_TOKEN is unset.")

    act = account_id(token)
    rows = collect(token, act)
    serving = [r for r in rows if r["serving"] == "yes"]

    print(f"account {act} -- {len(rows)} ad(s), {len(serving)} serving\n")
    print(f"{'utm_content':26} {'via':9} {'media':6} {'ad':34} status")
    for r in rows:
        mark = " " if r["serving"] == "yes" else "."
        print(f"{mark}{(r['utm_content'] or '(none)'):25} {r['tagged_via']:9} "
              f"{r['media']:6} {r['ad'][:34]:34} {r['status']}")

    # Two ads sharing a utm_content is the exact failure this sheet exists to
    # surface, so it is reported loudly rather than left to be noticed.
    dupes = {}
    for r in serving:
        dupes.setdefault(r["utm_content"] or "(none)", []).append(r["ad"])
    clashes = {k: v for k, v in dupes.items() if len(v) > 1}
    if clashes:
        print("\nDUPLICATE utm_content among SERVING ads:")
        for k, v in clashes.items():
            print(f"  {k} -> {', '.join(v)}")
    else:
        print("\nutm_content distinct across all serving ads.")

    # An ad tagged in BOTH places sends two utm_source values per click. It is
    # not a reporting gap -- it is a live defect, and the sheet is where it is
    # cheapest to notice. lint-ad-copy.js raises the same thing as an error.
    both = [r for r in serving if r["tagged_via"] == "both"]
    if both:
        print("\nTAGGED TWICE -- url_tags AND link params, so each sends two utm_source values:")
        for r in both:
            print(f"  {r['ad']}")

    untagged = [r for r in serving if r["tagged_via"] == "none"]
    if untagged:
        print("\nUNTAGGED among SERVING ads -- these record as \"direct\" in Stripe:")
        for r in untagged:
            print(f"  {r['ad']}")

    if args.dry_run:
        print("\nDry run -- workbook not written.")
        return
    # The generated workbook is created on first run -- it is script-owned, so
    # a missing file is a first run, not an error. Its FOLDER must exist,
    # because a missing folder means the path is wrong rather than new.
    folder = os.path.dirname(args.workbook)
    if folder and not os.path.isdir(folder):
        raise SystemExit(f"Folder not found: {folder}")
    fresh = not os.path.exists(args.workbook)
    write_sheet(args.workbook, rows)
    print(f"\n{'Created' if fresh else 'Wrote'} '{SHEET}' ({len(rows)} rows) "
          f"in {os.path.basename(args.workbook)}")


if __name__ == "__main__":
    main()
